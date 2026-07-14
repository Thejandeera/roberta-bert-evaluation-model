import os
import warnings
import pandas as pd
from transformers import pipeline
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Suppress standard Hugging Face warnings for a cleaner terminal output
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
warnings.filterwarnings("ignore")

print("Loading AI Models into memory... (This will take a few seconds)")
# Load the 28-emotion RoBERTa model
roberta_model = pipeline('text-classification', model='SamLowe/roberta-base-go_emotions')
# Load the 6-emotion BERT model
bert_model = pipeline('text-classification', model='bhadresh-savani/bert-base-uncased-emotion')
print("Models loaded successfully!\n")

def process_csv(input_filename="input.csv", output_filename="evaluation_results.xlsx"):
    print("=" * 60)
    print(" Business Call Emotion Analyzer - Bulk Processor")
    print("=" * 60)

    if not os.path.exists(input_filename):
        print(f"Error: Could not find '{input_filename}' in the current directory.")
        return

    print(f"\nReading data from {input_filename}...")
    df = pd.read_csv(input_filename)

    # Standardize column headers to lowercase to avoid case-sensitivity errors
    df.columns = [col.strip().lower() for col in df.columns]

    if 'text' not in df.columns or 'actual_emotion' not in df.columns:
        print("Error: The CSV must contain exactly 'text' and 'actual_emotion' columns.")
        return

    print("\nWhich model should process this dataset?")
    print("  [1] RoBERTa (28 fine-grained business emotions)")
    print("  [2] BERT (6 standard emotions)")
    
    model_choice = input("> Enter 1 or 2: ").strip()

    if model_choice == '1':
        print("\nRunning RoBERTa inference across the dataset...")
        model = roberta_model
        prediction_column = 'roberta_prediction'
    elif model_choice == '2':
        print("\nRunning BERT inference across the dataset...")
        model = bert_model
        prediction_column = 'bert_prediction'
    else:
        print("Invalid choice. Shutting down analyzer.")
        return

    # Process each row
    predictions = []
    for text in df['text']:
        try:
            result = model(str(text))[0]['label']
        except Exception:
            result = "error"
        predictions.append(result)

    # Add the predictions to the dataframe
    df[prediction_column] = predictions

    print(f"Applying color highlights and saving to {output_filename}...")
    # Export to Excel so we can apply the cell colors
    df.to_excel(output_filename, index=False)

    # Load the workbook to apply the formatting
    wb = load_workbook(output_filename)
    ws = wb.active

    # Define the highlight colors
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # Find the dynamic column indexes
    headers = [cell.value for cell in ws[1]]
    actual_col_idx = headers.index('actual_emotion') + 1
    pred_col_idx = headers.index(prediction_column) + 1

    # Iterate through rows and apply color logic
    for row in range(2, ws.max_row + 1):
        actual_val = str(ws.cell(row=row, column=actual_col_idx).value).strip().lower()
        pred_val = str(ws.cell(row=row, column=pred_col_idx).value).strip().lower()

        if actual_val == pred_val:
            ws.cell(row=row, column=pred_col_idx).fill = yellow_fill
        else:
            ws.cell(row=row, column=pred_col_idx).fill = red_fill

    # Save the final styled file
    wb.save(output_filename)
    print("-" * 60)
    print(f"Done! Results successfully saved with color highlights to {output_filename}")

if __name__ == "__main__":
    process_csv()